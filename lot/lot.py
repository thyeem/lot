import calendar
import math
import re
from collections import defaultdict
from datetime import datetime
from unicodedata import east_asian_width

import openpyxl
from openpyxl.styles import Font, PatternFill
from ortools.sat.python import cp_model

from ouch import *

from .parser import *

TEMPERATURE = 0.1
PENALTY_ENTROPY = 0.2
PENALTY_SIGMA = 0.2
CUT_ACTOR_NAME = 4


# ----------------------
# parse LOT source code
# ----------------------
def parse_lot(s):
    _, s = jump(s)
    g, s = parse_grid(s)
    _, s = parse_bar(s)
    a, s = parse_policy(s)
    if s:
        fail("Syntax Error: Invalid syntax used", s)
    return (g, dict(a)), s  # (grid, policy)


def parse_grid(s):
    g, s = sepby(symbol("+"), some(kwd_list))(s)
    _, s = jump(s)
    return g, s


def parse_bar(s):
    _, s = atleast(3, char("-"))(s)
    _, s = jump(s)
    return None, s


def parse_policy(s):
    def acts(s):
        a, s = symbol("@")(s)
        r, s = token(digits)(s)
        _, s = jump(s)
        return (a, r), s  # (@, digits)

    def rest(s):
        a, s = symbol("/")(s)
        r, s = token(digits)(s)
        _, s = jump(s)
        return (a, r), s  # (/, digits)

    def actor(s):
        return some(noneof("#<>\n"), fold=True)(s)

    def unit(s):
        a, s = token(angles(actor))(s)
        _, s = jump(s)
        q, s = many(choice(acts, rest))(s)
        r, s = many(choice(parse_x, parse_o, parse_excl))(s)
        r.extend(q)
        return (a, r), s  # (actor, preferences)

    return some(unit)(s)


def parse_x(s):
    _, s = symbol("-")(s)
    _, s = choice(symbol("X"), symbol("x"))(s)
    r, s = squares(sepby(symbol(","), xkwd))(s)
    return ("x", normalize(r)), s  # (x, [xkwd,...])


def parse_o(s):
    _, s = symbol("-")(s)
    _, s = choice(symbol("O"), symbol("o"))(s)
    r, s = squares(sepby(symbol(","), choice(rexpr, xkwd)))(s)
    return ("o", normalize(r)), s  # (o, [xkwd,...])


def parse_excl(s):
    _, s = symbol("-")(s)
    _, s = symbol("!")(s)
    r, s = squares(sepby(symbol(","), xkwd))(s)
    return ("!", normalize(r)), s  # (!, [xkwd,...])


def kwd(s):
    """
    >>> kwd("1-11;3")
    (['1', '4', '7', '10'], '')

    >>> kwd("keyword")
    (['keyword'], '')
    """
    r, s = some(
        label("'keyword-char'", noneof("[](),<=>:# \t\n")),
        fold=True,
    )(s)
    x = expand(r)
    if x:
        return x, s
    else:
        fail("Invalid keyword expression", s)


def kwd_list(s):
    """
    >>> kwd_list("[1-3, 4, 5-11;3]")
    (['1', '2', '3', '4', '5', '8', '11'], '')
    """
    r, s = squares(sepby(symbol(","), strip(kwd)))(s)
    return nub(flatten(r)), s


def kwd_tuple(s):
    """
    >>> kwd_tuple("(1-3, 4, 5-11;3)")
    (['1', '2', '3', '4', '5', '8', '11'], '')

    >>> kwd_tuple("((1-3, 4, (5-11;3)))")
    (['1', '2', '3', '4', '5', '8', '11'], '')
    """
    r, s = between(
        char("("),
        char(")"),
        sepby(symbol(","), choice(strip(kwd_tuple), strip(kwd))),
    )(s)
    return nub(flatten(r)), s


def xkwd(s):
    """
    >>> xkwd("sofia")
    ([['sofia']], '')

    >>> xkwd("(1-31;14):May:2025")
    ([['1', '15', '29'], ['May'], ['2025']], '')

    >>> xkwd("((1-31;14):May:2025)")
    ([['1', '15', '29'], ['May'], ['2025']], '')
    """

    def atom(s):
        return sepby(char(":"), choice(kwd_tuple, kwd))(s)

    r, s = choice(atom, parens(xkwd))(s)
    return r, s


def rexpr(s):
    """
    # >>> rexpr("sofia >= 10")
    # (('#', [(['sofia'],), '>=', '10']), '')

    # >>> rexpr("(1-31;14):May:2025 > 5")
    # (('#', [(['1', '15', '29'], ['May'], ['2025']), '>', '5']), '')
    """
    o = []
    r, s = token(xkwd)(s)
    o.append(tuple(r))
    r, s = choice(
        symbol("<="),
        symbol(">="),
        symbol("="),
        symbol("<"),
        symbol(">"),
    )(s)
    o.append(r)
    r, s = token(digits)(s)
    o.append(r)
    return ("#", o), s


def expand(x):
    """
    >>> expand("3")
    ['3']

    >>> expand("1-5")
    ['1', '2', '3', '4', '5']

    >>> expand("2-10;3")
    ['2', '5', '8']

    >>> expand("5-1")
    []
    """
    span = re.fullmatch(r"(\d+)\s*-\s*(\d+)", x)
    if span:
        return mapl(str, seq(*map(int, span.groups())))
    step = re.fullmatch(r"(\d+)\s*-\s*(\d+)\s*;\s*(\d+)", x)
    if step:
        i, j, k = map(int, step.groups())
        return mapl(str, seq(i, i + k, j))
    return [x]


def normalize(r):
    """
    >>> normalize([[['1','15','29'], ['May'], ['2025']]])
    [('1', 'May', '2025'), ('15', 'May', '2025'), ('29', 'May', '2025')]

    >>> normalize([('#', [(['maria'], ['yoajung']), '>', '3'])])
    [('#', (('maria', 'yoajung'), '>', '3'))]
    """
    o = []
    for x in r:
        if fst(x) == "#":
            key, sym, val = snd(x)
            o.append(("#", (fst(cartprodl(*key)), sym, val)))
        else:
            o.extend(cartprodl(*x))
    return o


# ---------------------------------------------------------------------
#        LOT | grid + --- + policy
# ---------------------------------------------------------------------
#      nodes | cartprod(grid)            ;; junctions of grid
#     policy | { actor: [preference] }   ;; actor's requests
#     actors | policy.keys()             ;; target group
#        act | (actor, *node)            ;; unique id in vars
#       vars | { act: model.boolvar }    ;; model variables
#     coeffs | { act: float }            ;; weights of availability
#  objective | sum(coeffs[act] * vars[act], ...)
# ---------------------------------------------------------------------


def solve(args):
    s = reader(args.FILE, "r", encoding="utf-8").read()
    (grid, policy), _ = parse_lot(s)
    validate_policy(grid, policy)
    nodes = gen_nodes(grid)
    rmap = gen_rmap(nodes)
    actors = policy.keys()
    rest = {actor: args.min_rest or 0 for actor in actors}
    consts = dict(
        grid=grid,
        policy=policy,
        nodes=nodes,
        actors=actors,
        rmap=rmap,
        rest=rest,
    )
    max_acts = len(nodes) // len(actors) or 1
    max_it = args.max_it or 5
    it = 0
    while True:
        model = cp_model.CpModel()
        vars = gen_vars(model, consts)
        coeffs = process_policy(model, vars, consts)
        rule_single_actor_per_node(model, vars, consts)
        rule_at_most_one_act_per_root(model, vars, consts)
        rule_clip_acts_per_actor(model, vars, consts, max_acts)
        rule_rest_between_acts(model, vars, consts)

        solver = cp_model.CpSolver()
        set_objective(model, vars, coeffs, consts)
        if solver.solve(model) in (cp_model.FEASIBLE, cp_model.OPTIMAL):
            o = collect_results(solver, vars, coeffs, consts)
            if valid_results(consts, o):
                report_and_export(consts, o, args)
                break
            else:
                continue
        if it > max_it:
            error(f"Aborted. Exceeded maximum iterations: {max_it}", e=Exception)
        max_acts += 1
        it += 1


def gen_nodes(grid):
    return cf_(dsort, concat)(cartprod(*g) for g in grid)


def gen_rmap(nodes):
    rmap = {}
    for node in nodes:
        r = fst(node)
        if r in rmap:
            rmap[r].append(node)
        else:
            rmap[r] = [node]
    return rmap


def gen_vars(model, consts):
    vars = {}
    for actor in consts["actors"]:
        for node in consts["nodes"]:
            act = (actor, *node)
            vars[act] = model.new_bool_var("_".join(act))
    return vars


def match_node(prefs, node):
    return any(set(o).issubset(set(node)) for o in prefs)


def dsort(x):
    return sort(
        x,
        key=cf_(
            lambda o: int(o) if o.isdigit() else o,
            op.itemgetter(0),
        ),
    )


def expr(sym, lhs, rhs):
    bop = {
        "<": op.lt,
        "<=": op.le,
        "=": op.eq,
        ">": op.gt,
        ">=": op.ge,
    }.get(sym)
    if not bop:
        error(f"Error, no such operator provided: {sym}", e=Exception)
    return bop(lhs, rhs)


def valid_results(consts, o):
    for actor in consts["actors"]:
        if not len([node for node in o["actors"][actor]]):
            return False
    return True


def validate_policy(grid, policy):
    keys = set(flat(grid))
    found = []
    for actor, prefs in policy.items():
        d = read_prefs(prefs)
        for k in flat(d["o"], d["x"], d["!"]):
            if k not in keys:
                found.append((k, actor))
    if found:
        out = ["Invalid keywords", ""]
        for k, actor in found:
            out.append(f"    '{k}', from <{actor}>")
        out.append("")
        out.append(f"Found {len(found)} error(s).")
        error(unlines(out), e=Exception)


def process_policy(model, vars, consts):
    coeffs = {}
    for actor, prefs in consts["policy"].items():
        d = read_prefs(prefs)
        # process @acts if any
        if d["@"]:
            model.add(
                sum(vars[(actor, *node)] for node in consts["nodes"]) == d["@"],
            )
        # update /rest if any
        if not isinstance(d["/"], list):
            consts["rest"][actor] = d["/"]
        # process q-preference
        for key, sym, val in d["q"]:
            lhs = sum(
                vars[(actor, *node)]
                for node in consts["nodes"]
                if match_node([key], node)
            )
            model.add(expr(sym, lhs, int(val)))
        # process o/x preference
        for node in consts["nodes"]:
            act = (actor, *node)
            coeffs[act] = 1 if not d["o"] or match_node(d["o"], node) else 0
            if match_node(d["x"], node):
                model.add(vars[act] == 0)
        # update priority
        if d["!"]:
            precedence = zipl(d["!"], w_priority(len(d["!"])))
            for el, w in precedence:
                for node in consts["nodes"]:
                    if set(el).issubset(set(node)):
                        act = (actor, *node)
                        if coeffs[act]:
                            coeffs[act] += w
    return coeffs


def w_priority(n):
    if n < 1:
        return []
    norm = 1 / math.sqrt(n)
    return [-math.log10(i / int(n)) * norm for i in seq(1, n)]


def read_prefs(prefs):
    d = defaultdict(list)
    for sym, o in prefs:
        if sym == "o":
            for el in o:
                if fst(el) == "#":
                    d["o"].append(fst(snd(el)))
                    d["q"].append(snd(el))
                else:
                    d["o"].append(el)
        elif sym == "x":
            d[sym].extend(o)
        elif sym == "!":
            d[sym] = o
        elif sym in ["@", "/"]:
            d[sym] = int(o)
        else:
            error(f"Error, used illegal symbol: {sym}", e=Exception)
    return d


def rule_single_actor_per_node(model, vars, consts):
    for node in consts["nodes"]:
        model.add(sum(vars[(actor, *node)] for actor in consts["actors"]) == 1)


def rule_at_most_one_act_per_root(model, vars, consts):
    root = cf_(uniq, fst, zip)(*consts["nodes"])
    for actor in consts["actors"]:
        for r in root:
            model.add(sum(vars[(actor, *node)] for node in consts["rmap"][r]) <= 1)


def rule_clip_acts_per_actor(model, vars, consts, max_acts):
    for actor in consts["actors"]:
        acts = []
        for node in consts["nodes"]:
            act = (actor, *node)
            acts.append(vars[act])
        model.add(1 <= sum(acts))  # assign at least once per actor
        model.add(sum(acts) <= max_acts)


def rule_rest_between_acts(model, vars, consts):
    root = cf_(uniq, fst, zip)(*consts["nodes"])
    for actor in consts["actors"]:
        min_rest = consts["rest"][actor]
        if not min_rest:
            continue
        for i, r in enumerate(root[:-min_rest]):
            sched = model.new_bool_var(f"sched_{actor}_{r}")
            model.add_bool_or(
                [vars[(actor, *node)] for node in consts["rmap"][r]]
            ).only_enforce_if(sched)
            model.add_bool_and(
                [vars[(actor, *node)].Not() for node in consts["rmap"][r]]
            ).only_enforce_if(sched.Not())
            for r_ in root[i + 1 : i + 1 + min_rest]:
                for node in consts["rmap"][r_]:
                    model.add(vars[(actor, *node)] == 0).only_enforce_if(sched)


def set_objective(model, vars, coeffs, consts):
    def add_noise(x):
        return x + rand(TEMPERATURE) if x >= 1 else x

    penalty_entropy = penalize_low_entropy(model, vars, consts)
    penalty_sigma = penalize_high_sigma(model, vars, consts)
    model.maximize(
        sum(
            add_noise(coeffs[(actor, *node)]) * vars[(actor, *node)]
            for actor in consts["actors"]
            for node in consts["nodes"]
        )
        - penalty_entropy
        - penalty_sigma
    )


def penalize_low_entropy(model, vars, consts):
    penalties = []
    d = defaultdict(list)
    for node in consts["nodes"]:
        for e in node:
            d[e].append(node)
    d = {k: v for k, v in d.items() if len(v) > 1}
    for actor in consts["actors"]:
        for _, nodes in d.items():
            el_vars = [vars[(actor, *node)] for node in nodes]
            penalty = model.new_int_var(0, len(el_vars) - 1, "")
            model.add(penalty >= sum(el_vars) - 1)
            penalties.append(penalty)
    return PENALTY_ENTROPY * sum(penalties)


def penalize_high_sigma(model, vars, consts):
    penalties = []
    num_nodes = len(consts["nodes"])
    for actor in consts["actors"]:
        penalty = model.new_int_var(0, num_nodes**2, "")
        sum_acts = model.new_int_var(0, num_nodes, "")
        acts = [vars[(actor, *node)] for node in consts["nodes"]]
        model.add(sum_acts == sum(acts))
        model.add_multiplication_equality(penalty, [sum_acts, sum_acts])
        penalties.append(penalty)
    return PENALTY_SIGMA * sum(penalties)


def collect_results(solver, vars, coeffs, consts):
    o = dict(nodes={}, actors=defaultdict(list))
    for actor in consts["actors"]:
        for node in consts["nodes"]:
            act = (actor, *node)
            if solver.value(vars[act]) == 1:
                if coeffs[act]:
                    o["nodes"][node] = actor
                    o["actors"][actor].append(node)
                else:
                    o["nodes"][node] = "*"
    return o


# ----------------------
# report and dump files
# ----------------------
def report_and_export(consts, o, args):
    def jx(node):
        return ":".join(node)

    def cut(x):
        return x.split()[0][:CUT_ACTOR_NAME]

    def uni(s):
        return sum(2 if east_asian_width(c) in {"W", "F"} else 1 for c in s)

    def by_actors():
        print()
        for actor in sort(consts["actors"]):
            print(actor)
            print("\n".join([" " * 8 + jx(node) for node in o["actors"][actor]]))

    def by_nodes():
        c = max(uni(jx(node)) for node in consts["nodes"]) + 4
        print()
        for node in consts["nodes"]:
            print(f"{justf(jx(node), c, pad='.')}", f"{o['nodes'][node]}")

    def base_cal():
        year = args.year or datetime.today().year
        month = args.month or datetime.today().month
        cal = calendar.Calendar(firstweekday=6).monthdayscalendar(year, month)
        days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]
        events = defaultdict(list)
        for node in consts["nodes"]:
            events[int(fst(node))].append(f"{last(node)}:{cut(o['nodes'][node])}")
        return cal, days, events, year, month

    def to_cal():
        cal, days, events, *_ = base_cal()
        c = max(uni(x) for e in events.values() for x in e)
        bar = "-" * 7 * (c + 3)
        nul = " " * c
        print()
        print(bar)
        print(" | ".join(justf(str(day), c, "^") for day in days))
        print(bar)
        for week in cal:
            row = []
            for day in week:
                row.append(nul if day == 0 else justf(f"{day}", c))
            print(" | ".join(row))
            for i in range(max(len(events.get(day, [])) for day in week)):
                e = []
                for day in week:
                    e.append(
                        nul
                        if day == 0 or i >= len(events.get(day, []))
                        else justf(f"{events[day][i]}", c)
                    )
                print(" | ".join(e))
            print(bar)

    def adjust(ws):
        for col in ws.iter_cols():
            c = max(uni(f"{cell.value}") if cell.value else 0 for cell in col)
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col[0].column)
            ].width = (c + 2)

    def by_actors_xl(wb):
        ws = wb.create_sheet(title="Actors")
        for actor in sort(consts["actors"]):
            for i, node in enumerate(o["actors"][actor]):
                ws.append(["", jx(node)] if i else [actor, jx(node)])
        adjust(ws)

    def by_nodes_xl(wb):
        ws = wb.create_sheet(title="Nodes")
        for node in consts["nodes"]:
            ws.append([jx(node), o["nodes"][node]])
        adjust(ws)

    def to_cal_xl(wb):
        cal, days, events, year, month = base_cal()
        ws = wb.create_sheet(title=f"{year}-{month:02d}")
        skyblue = PatternFill(start_color="cfecf7", fill_type="solid")
        gray = PatternFill(start_color="efefef", fill_type="solid")
        bold = Font(bold=True)
        for col, day in enumerate(days, start=1):
            cell = ws.cell(row=1, column=col, value=day)
            cell.fill = gray
            cell.font = bold
        for week in cal:
            row = []
            for col, day in enumerate(week, start=1):
                row.append("" if day == 0 else day)
            ws.append(row)
            for col, day in enumerate(week, start=1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.fill = skyblue
                cell.font = bold
            for i in range(max(len(events.get(day, [])) for day in week)):
                e = []
                for day in week:
                    e.append(
                        ""
                        if day == 0 or i >= len(events.get(day, []))
                        else events[day][i]
                    )
                ws.append(e)
        adjust(ws)

    if args.actor:
        by_actors()
    if args.node:
        by_nodes()
    if args.cal:
        to_cal()
    if args.output:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        to_cal_xl(wb)
        by_nodes_xl(wb)
        by_actors_xl(wb)
        wb.save(
            args.output
            if args.output.split(".")[-1] in ("xlsx", "xls")
            else f"{args.output}.xlsx"
        )
